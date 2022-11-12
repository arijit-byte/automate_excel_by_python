from google_images_download import google_images_download
import openpyxl
import os
from openpyxl.drawing.image import Image
from openpyxl import workbook, load_workbook
import pandas as pd
response = google_images_download.googleimagesdownload()


df = pd.read_excel("C:\python_project\Image_Test_original.xlsx")                 #  ===================>   1) SOURCE EXCEL FILE = USER I/P

# 2) ===================> DOWNLOAD IMAGES TO FOLDER = USER I/P

for i, k in zip(df["Type"],df["Company"]):
    query_var = k + " company logo"
    arguments = {"output_directory":"C:\\python_project\\ALL_IMAGES_FOLDER\\", "keywords":query_var, "limit":1, "print_urls":False, "silent_mode":True, "format":"png"}
    response.download(arguments)

# -------------------------------------------------2nd part

wb = load_workbook("C:\\python_project\\Image_Test_original.xlsx")     # 3) ===================> SOURCE EXCEL FILE = USER I/P, same as (1)
ws = wb.active
# ------------------------Taking from local folder
img_dir = "C:\\python_project\\ALL_IMAGES_FOLDER"                      #  4) ===================> DOWNLOAD IMAGES FROM FOLDER = USER I/P, same as (2)
img_folder = os.listdir(img_dir)
img_crop = [_.split(" ")[0] for _ in img_folder]

# ------------------------Taking from excel

for r in range(2,ws.max_row+1):
    my_row_cnt = r
    for c in range(2,ws.max_column+1):
        comp = (ws.cell(row=r, column=c).value)
        if comp in img_crop:
            fetch_index_img_crop = img_crop.index(comp)
            img_path_from_local = img_folder[fetch_index_img_crop]
            fecth_exact_img = img_dir + '\\' + img_path_from_local

            # insert fecth_exact_img into excel cell where company == comp
            for imgs in os.listdir(fecth_exact_img):
                print("Just Entered")
                temp_path_builder = fecth_exact_img + "\\" + imgs
                print("s_1")

                try:
                    im = Image(temp_path_builder)
                    print("s_2")
                    im.height = 20
                    im.width = 20
                    print("s_3")
                    ws.add_image(im, anchor='C'+str(my_row_cnt))
                    print("s_4")
                except:
                    print("Image Format not correct")

#         # comp_names.append(comp)

wb.save("C:\\python_project\\FINAL_EXCEL.xlsx")                 #  5) Finally save the imgs here = USER I/P,
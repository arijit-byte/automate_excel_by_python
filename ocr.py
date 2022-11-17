import pytesseract
import cv2
import re
import os
import pandas as pd
from openpyxl.drawing.image import Image
from openpyxl_image_loader import SheetImageLoader
from openpyxl import workbook, load_workbook
#
print("1")
pytesseract.pytesseract.tesseract_cmd=r'C:\ocr_proj_new\ocr_pack\tesseract.exe'
print("2")
wb = load_workbook("FINAL_EXCEL_1.xlsx")
print("3")
ws = wb.active
print("4")

df  = pd.read_excel("FINAL_EXCEL_1.xlsx")
print("5")
wimg = wb["Sheet1"]
print("6")
image_loder = SheetImageLoader(wimg)
print("7")
print("###############################################################################################")
for i, r in df.iterrows():
    print("FOR -> 1")
    #----------------IMG
    im_index = "C" + str(i + 2)
    print("FOR -> 2")
    image = image_loder.get(im_index)
    print("FOR -> 3")
    text = pytesseract.image_to_string(image)
    print("FOR -> 4")
    text_splited = text.split('\n')   # Here is the prob Return splited texts in list
    print("FOR -> 5")
    text_joined = " ".join(text_splited)
    print("FOR -> 6")
    res = re.sub(' +', ' ', text_joined)
    print("FOR -> 7")
    # ----------------IMG
    insert_index = "D" + str(i + 2)
    print("FOR -> 8")
    ws[insert_index] = res
    print("FOR -> 9")


print("OUT")
wb.save(r"FINAL_EXCEL_1.xlsx")
# wb.close()











#-------------------------OCR PART
# img_ocr = cv2.imread("PWC.png")
#
# text = pytesseract.image_to_string(img_ocr)
# text_splited = text.split('\n')  # Return splited texts in list
# text_joined = " ".join(text_splited)
# res = re.sub(' +', ' ', text_joined)
#
# print(res)
#-------------------------OCR PART



















#---------------------COMMENT----------------------------------------------
# for r in range(2,ws.max_row+1):
#     my_row_cnt = r
#     for c in range(3, ws.max_column+1):
#         image = image_loder.get(ws.cell(row=r, column=c))
#
#         print(ws.cell(row=r, column=c))
#         image.show()
#----------------------COMMENT---------------------------------------------
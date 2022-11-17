import pytesseract
import cv2
import re
import os
import pandas as pd
from openpyxl.drawing.image import Image
from openpyxl_image_loader import SheetImageLoader
from openpyxl import workbook, load_workbook
#
pytesseract.pytesseract.tesseract_cmd=r'C:\ocr_proj_new\ocr_pack\tesseract.exe'
wb = load_workbook("FINAL_EXCEL_1.xlsx")
ws = wb.active
#------------------------DF
df  = pd.read_excel("FINAL_EXCEL_1.xlsx")
wimg = wb["Sheet1"]
image_loder = SheetImageLoader(wimg)
#------------------------DF
for i, r in df.iterrows():
    print("1>>>>>>>>>>",i)
    res = "LAST"
    insert_index = "D" + str(i + 2)
    print("1-----------",insert_index)
    ws[insert_index] = res

wb.save(r"FINAL_EXCEL_1.xlsx")



# for i, r in df.iterrows():
#     im_index = "C" + str(i + 2)
#     image = image_loder.get(im_index)
#     text = pytesseract.image_to_string(image)
#     text_splited = text.split('\n')            # Return splited texts in list
#     text_joined = " ".join(text_splited)
#     res = re.sub(' +', ' ', text_joined)
#     print(res)
#     insert_index = "D" + str(i + 2)
#     ws[insert_index] = res

# wb.save(r"FINAL_EXCEL_1.xlsx")
    # break






#-------------------------OCR PART
# img_ocr = cv2.imread("PWC.png")
#
# text = pytesseract.image_to_string(img_ocr)
# text_splited = text.split('\n')  # Return splited texts in list
# text_joined = " ".join(text_splited)
# res = re.sub(' +', ' ', text_joined)
#
# print(res)
#-------------------------OCR PART



















#---------------------COMMENT----------------------------------------------
# for r in range(2,ws.max_row+1):
#     my_row_cnt = r
#     for c in range(3, ws.max_column+1):
#         image = image_loder.get(ws.cell(row=r, column=c))
#
#         print(ws.cell(row=r, column=c))
#         image.show()
#----------------------COMMENT---------------------------------------------